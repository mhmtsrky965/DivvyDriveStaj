import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from datetime import datetime
import threading
import re
import getpass
import string
import platform

try:
    from docx import Document
    import pdfplumber
    import openpyxl
    from openpyxl import load_workbook


    if platform.system() == "Windows":
        try:
            import win32security
            import win32api
            WINDOWS_SECURITY_AVAILABLE = True
        except ImportError:
            WINDOWS_SECURITY_AVAILABLE = False
    else:
        WINDOWS_SECURITY_AVAILABLE = False

except ImportError as e:
    print(f"Gerekli kütüphaneler yüklenmemiş: {e}")
    print("pip install python-docx pdfplumber openpyxl")
    if platform.system() == "Windows":
        print("Windows için: pip install pywin32")

class FileDetailWindow:
    def __init__(self, parent, file_info, search_term="", search_type="İçeren", case_sensitive=False):
        self.window = tk.Toplevel(parent)
        self.window.title(f"Dosya Detayı - {file_info['file_name']}")
        self.window.geometry("800x600")
        self.window.transient(parent)
        self.window.grab_set()

        self.file_info = file_info
        self.search_term = search_term if not case_sensitive else search_term
        self.search_type = search_type
        self.case_sensitive = case_sensitive

        self.setup_ui()
        self.display_content()

    def setup_ui(self):
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)


        info_frame = ttk.LabelFrame(main_frame, text="Dosya Bilgileri", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))

        info_labels = [
            ("Dosya Adı:", self.file_info['file_name']),
            ("Uzantı:", self.file_info['extension']),
            ("Boyut:", self.file_info['size']),
            ("Yazar:", self.file_info['author']),
            ("Belge Tarihi:", self.file_info['document_date']),
            ("Oluşturma Tarihi:", self.file_info['creation_date']),
            ("Sahip:", self.file_info['owner'])
        ]

        for i, (label, value) in enumerate(info_labels):
            row = i // 2
            col = (i % 2) * 2
            ttk.Label(info_frame, text=label, font=('Arial', 9, 'bold')).grid(
                row=row, column=col, sticky=tk.W, padx=(0, 5), pady=2)
            ttk.Label(info_frame, text=value).grid(
                row=row, column=col+1, sticky=tk.W, padx=(0, 20), pady=2)


        content_frame = ttk.LabelFrame(main_frame, text="Dosya İçeriği", padding="10")
        content_frame.pack(fill=tk.BOTH, expand=True)

        self.text_widget = tk.Text(content_frame, wrap=tk.WORD, font=('Arial', 10))
        scrollbar = ttk.Scrollbar(content_frame, orient=tk.VERTICAL, command=self.text_widget.yview)
        self.text_widget.configure(yscrollcommand=scrollbar.set)

        self.text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)


        self.text_widget.tag_config("highlight", background="yellow", foreground="red", font=('Arial', 10, 'bold'))
        self.text_widget.tag_config("normal", font=('Arial', 10))

        ttk.Button(main_frame, text="Kapat", command=self.window.destroy).pack(pady=(10, 0))

    def display_content(self):
        content = self.file_info['content']

        if self.search_term:
            self.highlight_search_terms(content)
        else:
            self.text_widget.insert(tk.END, content, "normal")

        self.text_widget.config(state=tk.DISABLED)

    def clean_word(self, word):
        """Kelimeden noktalama işaretlerini temizler"""
        return word.strip(string.punctuation)

    def highlight_search_terms(self, content):
        if not self.search_term:
            self.text_widget.insert(tk.END, content, "normal")
            return

        if self.search_type == "İçeren":
            if self.case_sensitive:
                pattern = re.compile(re.escape(self.search_term))
            else:
                pattern = re.compile(re.escape(self.search_term), re.IGNORECASE)

            last_end = 0
            for match in pattern.finditer(content):
                if match.start() > last_end:
                    self.text_widget.insert(tk.END, content[last_end:match.start()], "normal")
                self.text_widget.insert(tk.END, content[match.start():match.end()], "highlight")
                last_end = match.end()

            if last_end < len(content):
                self.text_widget.insert(tk.END, content[last_end:], "normal")

        else:
            if self.search_type == "Başlayan":
                pattern = r'\b' + re.escape(self.search_term)
            else:
                pattern = re.escape(self.search_term) + r'(?=\s|[^\w]|$)'

            if self.case_sensitive:
                regex = re.compile(pattern)
            else:
                regex = re.compile(pattern, re.IGNORECASE)

            last_end = 0
            for match in regex.finditer(content):
                if match.start() > last_end:
                    self.text_widget.insert(tk.END, content[last_end:match.start()], "normal")
                self.text_widget.insert(tk.END, content[match.start():match.end()], "highlight")
                last_end = match.end()

            if last_end < len(content):
                self.text_widget.insert(tk.END, content[last_end:], "normal")

class FileManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ODEV2")
        self.root.geometry("1400x800")

        self.file_data = []
        self.current_search_term = ""
        self.current_search_type = "İçeren"
        self.current_case_sensitive = False

        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))


        upload_frame = ttk.LabelFrame(main_frame, text="Dosya Yükleme", padding="10")
        upload_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Button(upload_frame, text="Dosyaları Seç ve Yükle",
                  command=self.upload_files).grid(row=0, column=0, padx=(0, 10))

        self.progress = ttk.Progressbar(upload_frame, mode='indeterminate')
        self.progress.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))

        self.status_label = ttk.Label(upload_frame, text="Hazır")
        self.status_label.grid(row=1, column=0, columnspan=2, pady=(5, 0))


        search_frame = ttk.LabelFrame(main_frame, text="Arama", padding="10")
        search_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))


        ttk.Label(search_frame, text="Arama Türü:").grid(row=0, column=0, sticky=tk.W)
        self.search_type = tk.StringVar(value="İçeren")
        search_combo = ttk.Combobox(search_frame, textvariable=self.search_type,
                                   values=["İçeren", "Başlayan", "Biten"], state="readonly")
        search_combo.grid(row=0, column=1, padx=(5, 10), sticky=(tk.W, tk.E))

        ttk.Label(search_frame, text="Arama Terimi:").grid(row=0, column=2, sticky=tk.W)
        self.search_entry = ttk.Entry(search_frame, width=30)
        self.search_entry.grid(row=0, column=3, padx=(5, 10), sticky=(tk.W, tk.E))
        self.search_entry.bind('<Return>', lambda e: self.search_files())

        ttk.Button(search_frame, text="Ara", command=self.search_files).grid(row=0, column=4, padx=(5, 0))
        ttk.Button(search_frame, text="Temizle", command=self.clear_search).grid(row=0, column=5, padx=(5, 0))


        self.case_sensitive = tk.BooleanVar()
        case_check = ttk.Checkbutton(search_frame, text="Büyük-küçük harf duyarlı",
                                    variable=self.case_sensitive)
        case_check.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))


        table_frame = ttk.LabelFrame(main_frame, text="Dosya Listesi", padding="10")
        table_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))

        columns = ("Dosya Adı", "Uzantı", "Boyut", "Yazar", "Belge Tarihi", "Oluşturma Tarihi", "Sahip", "İşlem")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)

        column_widths = {"Dosya Adı": 160, "Uzantı": 60, "Boyut": 70, "Yazar": 110,
                "Belge Tarihi": 90, "Oluşturma Tarihi": 120, "Sahip": 110, "İşlem": 180}

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths.get(col, 120))

        self.tree.bind("<Double-1>", self.on_item_double_click)

        scrollbar_v = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_h = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)

        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        scrollbar_h.grid(row=1, column=0, sticky=(tk.W, tk.E))

        ttk.Button(table_frame, text="Detaylı Görüntüle",
                  command=self.show_file_detail).grid(row=2, column=0, pady=(10, 0), sticky=tk.W)


        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        search_frame.columnconfigure(3, weight=1)
        upload_frame.columnconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def find_author_in_content(self, content):
        """Dosya içeriğinde yazar bilgisini anahtar kelimelerle arar"""
        if not content:
            return "Bilinmiyor"


        search_content = content[:2000]


        author_keywords = [
            r'yazar[:\s]+([^\n\r]{1,100})',
            r'yazan[:\s]+([^\n\r]{1,100})',
            r'hazırlayan[:\s]+([^\n\r]{1,100})',
            r'hazırladen[:\s]+([^\n\r]{1,100})',
            r'yazarı[:\s]+([^\n\r]{1,100})',
            r'author[:\s]+([^\n\r]{1,100})',
            r'written by[:\s]+([^\n\r]{1,100})',
            r'prepared by[:\s]+([^\n\r]{1,100})',
            r'by[:\s]+([^\n\r]{1,100})',
            r'tarafından[:\s]*yazılan',
            r'([^\n\r]{1,50})\s+tarafından\s+yazılmıştır',
            r'([^\n\r]{1,50})\s+tarafından\s+hazırlanmıştır'
        ]

        for pattern in author_keywords:
            matches = re.finditer(pattern, search_content, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                if match.group(1):
                    author = match.group(1).strip()

                    if 2 <= len(author) <= 100:

                        if not re.match(r'^\d', author) and not author.isdigit():

                            author = re.sub(r'^[:\-\s]+', '', author)
                            author = re.sub(r'[:\-\s]+$', '', author)
                            if author:
                                return author

        return "Bilinmiyor"

    def find_date_in_content(self, content):
        """Dosya içeriğinde tarih bilgisini arar"""
        if not content:
            return "Bilinmiyor"


        search_content = content[:1500]

        date_patterns = [
            r'(\d{1,2}[./\-]\d{1,2}[./\-]\d{4})',
            r'(\d{4}[./\-]\d{1,2}[./\-]\d{1,2})',
            r'(\d{1,2}\s+(?:Ocak|Şubat|Mart|Nisan|Mayıs|Haziran|Temmuz|Ağustos|Eylül|Ekim|Kasım|Aralık)\s+\d{4})',
            r'(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})',
            r'(?:tarih|date)[:\s]+([^\n\r]{1,50})',
            r'(\d{4})\s*yılı',
            r'©\s*(\d{4})',
        ]

        for pattern in date_patterns:
            matches = re.finditer(pattern, search_content, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                date_str = match.group(1).strip()
                if date_str and len(date_str) >= 4:

                    if re.match(r'^\d{4}$', date_str):
                        if 1900 <= int(date_str) <= 2030:
                            return date_str
                    elif not date_str.isdigit():
                        return date_str

        return "Bilinmiyor"

    def find_owner_in_content(self, content):
        """Dosya içeriğinde sahip bilgisini anahtar kelimelerle arar"""
        if not content:
            return "Bilinmiyor"


        search_content = content[:2000]


        owner_keywords = [
            r'sahip[:\s]+([^\n\r]{1,100})',
            r'sahibi[:\s]+([^\n\r]{1,100})',
            r'owner[:\s]+([^\n\r]{1,100})',
            r'belongs to[:\s]+([^\n\r]{1,100})',
            r'property of[:\s]+([^\n\r]{1,100})',
            r'kurum[:\s]+([^\n\r]{1,100})',
            r'kuruluş[:\s]+([^\n\r]{1,100})',
            r'organization[:\s]+([^\n\r]{1,100})',
            r'company[:\s]+([^\n\r]{1,100})',
            r'şirket[:\s]+([^\n\r]{1,100})'
        ]

        for pattern in owner_keywords:
            matches = re.finditer(pattern, search_content, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                if match.group(1):
                    owner = match.group(1).strip()
                    if 2 <= len(owner) <= 100:
                        if not re.match(r'^\d', owner) and not owner.isdigit():
                            owner = re.sub(r'^[:\-\s]+', '', owner)
                            owner = re.sub(r'[:\-\s]+$', '', owner)
                            if owner:
                                return owner

        return "Bilinmiyor"

    def get_file_owner(self, file_path):
        """Dosya sahibini gerçek olarak almaya çalışır"""
        try:
            if platform.system() == "Windows" and WINDOWS_SECURITY_AVAILABLE:
                # Windows için gerçek dosya sahibi
                try:
                    sd = win32security.GetFileSecurity(file_path, win32security.OWNER_SECURITY_INFORMATION)
                    owner_sid = sd.GetSecurityDescriptorOwner()
                    name, domain, type = win32security.LookupAccountSid(None, owner_sid)
                    return f"{domain}\\{name}" if domain else name
                except Exception:
                    pass

            # Unix/Linux sistemler için
            if platform.system() in ["Linux", "Darwin"]:
                import pwd
                stat_info = os.stat(file_path)
                try:
                    owner = pwd.getpwuid(stat_info.st_uid).pw_name
                    return owner
                except KeyError:
                    return f"UID: {stat_info.st_uid}"

            # Varsayılan olarak mevcut kullanıcı
            return getpass.getuser()
        except Exception:
            return "Bilinmiyor"

    def upload_files(self):
        file_types = [
            ("Tüm Desteklenen", "*.docx;*.pdf;*.txt;*.xlsx;*.xls"),
            ("Word Dosyaları", "*.docx"),
            ("PDF Dosyaları", "*.pdf"),
            ("Text Dosyaları", "*.txt"),
            ("Excel Dosyaları", "*.xlsx;*.xls")
        ]

        files = filedialog.askopenfilenames(title="Yüklenecek dosyaları seçin", filetypes=file_types)

        if files:
            self.progress.start()
            self.status_label.config(text="Dosyalar işleniyor...")
            thread = threading.Thread(target=self.process_files, args=(files,))
            thread.daemon = True
            thread.start()

    def process_files(self, files):
        for file_path in files:
            try:
                file_info = self.extract_file_info(file_path)
                if file_info:
                    self.file_data.append(file_info)
                    self.root.after(0, self.update_table_with_file, file_info)
            except Exception as e:
                print(f"Dosya işleme hatası {file_path}: {e}")

        self.root.after(0, self.processing_complete)

    def processing_complete(self):
        self.progress.stop()
        self.status_label.config(text=f"Toplam {len(self.file_data)} dosya yüklendi")

    def extract_file_info(self, file_path):
        try:
            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()
            file_size = self.format_file_size(os.path.getsize(file_path))

            stat_info = os.stat(file_path)
            creation_time = datetime.fromtimestamp(stat_info.st_ctime).strftime("%Y-%m-%d %H:%M")

            content = ""
            metadata_author = "Bilinmiyor"
            owner = self.get_file_owner(file_path)

            if file_ext == '.docx':
                content, metadata_author = self.read_docx(file_path)
            elif file_ext == '.pdf':
                content, metadata_author = self.read_pdf(file_path)
            elif file_ext == '.txt':
                content = self.read_txt(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                content, metadata_author = self.read_excel(file_path)


            content_author = self.find_author_in_content(content)
            final_author = content_author if content_author != "Bilinmiyor" else metadata_author

            document_date = self.find_date_in_content(content)

            content_owner = self.find_owner_in_content(content)
            final_owner = content_owner if content_owner != "Bilinmiyor" else owner

            return {
            'file_path': file_path,
            'file_name': file_name,
            'extension': file_ext,
            'size': file_size,
            'author': final_author,
            'document_date': document_date,
            'creation_date': creation_time,
            'owner': final_owner,
            'content': content
            }
        except Exception as e:
            print(f"Dosya bilgi çıkarma hatası: {e}")
            return None

    def read_docx(self, file_path):
        try:
            doc = Document(file_path)
            content = ""
            for paragraph in doc.paragraphs:
                content += paragraph.text + "\n"
            author = doc.core_properties.author or "Bilinmiyor"
            return content.strip(), author
        except Exception as e:
            return "", "Bilinmiyor"

    def read_pdf(self, file_path):
        try:
            content = ""
            author = "Bilinmiyor"
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        content += text + "\n"
                if pdf.metadata and pdf.metadata.get('Author'):
                    author = pdf.metadata.get('Author')
            return content.strip(), author
        except Exception as e:
            return "", "Bilinmiyor"

    def read_txt(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='cp1254') as file:
                    return file.read()
            except:
                return ""

    def read_excel(self, file_path):
        try:
            wb = load_workbook(file_path)
            content = ""
            author = "Bilinmiyor"

            try:
                if wb.properties and wb.properties.creator:
                    author = wb.properties.creator
                elif wb.properties and wb.properties.lastModifiedBy:
                    author = wb.properties.lastModifiedBy
            except:
                pass

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                content += f"Sayfa: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) for cell in row if cell is not None]
                    if row_text:
                        content += " | ".join(row_text) + "\n"
                content += "\n"
            return content.strip(), author
        except Exception as e:
            return "", "Bilinmiyor"

    def format_file_size(self, size_bytes):
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024**2:
            return f"{size_bytes/1024:.1f} KB"
        elif size_bytes < 1024**3:
            return f"{size_bytes/(1024**2):.1f} MB"
        else:
            return f"{size_bytes/(1024**3):.1f} GB"

    def clean_word(self, word):
        """Kelimeden noktalama işaretlerini temizler"""
        return word.strip(string.punctuation)

    def update_table_with_file(self, file_info):
        self.tree.insert("", "end", values=(
        file_info['file_name'],
        file_info['extension'],
        file_info['size'],
        file_info['author'],
        file_info['document_date'],
        file_info['creation_date'],
        file_info['owner'],
        "Detaylı görüntüleme için çift tıklayınız"
        ))

    def search_files(self):
        search_term = self.search_entry.get().strip()
        search_type = self.search_type.get()
        case_sensitive = self.case_sensitive.get()

        self.current_search_term = search_term
        self.current_search_type = search_type
        self.current_case_sensitive = case_sensitive

        if not search_term:
            self.show_all_files()
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        found_files = []

        for file_info in self.file_data:
            content = file_info['content']
            found = False

            if search_type == "İçeren":
                # İçeren arama - direkt string kontrolü
                if case_sensitive:
                    found = search_term in content
                else:
                    found = search_term.lower() in content.lower()

            elif search_type == "Başlayan":
                # Başlayan arama - kelime başında
                if case_sensitive:
                    pattern = r'\b' + re.escape(search_term)
                    found = bool(re.search(pattern, content))
                else:
                    pattern = r'\b' + re.escape(search_term)
                    found = bool(re.search(pattern, content, re.IGNORECASE))

            elif search_type == "Biten":
                # Biten arama - kelime sonunda veya noktalama işareti/boşluk öncesi
                escaped_term = re.escape(search_term)
                if case_sensitive:
                    # Aranan terim kelime sonunda veya noktalama/boşluk öncesi
                    pattern = escaped_term + r'(?=\s|[^\w]|$)'
                    found = bool(re.search(pattern, content))
                else:
                    pattern = escaped_term + r'(?=\s|[^\w]|$)'
                    found = bool(re.search(pattern, content, re.IGNORECASE))

            if found:
                found_files.append(file_info)

        for file_info in found_files:
            self.update_table_with_file(file_info)

        self.status_label.config(text=f"Arama sonucu: {len(found_files)} dosya bulundu")

    def clear_search(self):
        self.search_entry.delete(0, tk.END)
        self.case_sensitive.set(False)
        self.current_search_term = ""
        self.current_search_type = "içeren"
        self.current_case_sensitive = False
        self.show_all_files()

    def show_all_files(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for file_info in self.file_data:
            self.update_table_with_file(file_info)

        self.status_label.config(text=f"Toplam {len(self.file_data)} dosya gösteriliyor")

    def on_item_double_click(self, event):
        self.show_file_detail()

    def show_file_detail(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen bir dosya seçin!")
            return

        item = selection[0]
        file_name = self.tree.item(item, 'values')[0]

        file_info = None
        for info in self.file_data:
            if info['file_name'] == file_name:
                file_info = info
                break

        if file_info:
            FileDetailWindow(self.root, file_info, self.current_search_term,
                           self.current_search_type, self.current_case_sensitive)

if __name__ == "__main__":
    root = tk.Tk()
    app = FileManagerApp(root)
    root.mainloop()